from sklearn.decomposition import PCA
from sklearn.model_selection import cross_val_score
from sklearn.linear_model import LogisticRegression
import openpyxl
import matplotlib.pyplot as plt

# Importing data
features = 'features'    # Change to the dataset with the features
dataset = r'train.xlsx'
wb_train = openpyxl.load_workbook(train)
ws_train = wb_train.active
y_train = [ws_train.cell(row=i,column=2).value for i in range(2,ws_train.max_row)]

# Cross-validation
'''This section can be used to determine the number of components that
should be used for the PCA. The number of components that maximizes
performance, in this case accuracy, should be selected.''' 
'https://scikit-learn.org/1.5/modules/cross_validation.html'
pca = PCA(n_components=1)  # Initial number of components
scores = []
for n in range(1, pca.n_components_ + 1):
    pca_n = PCA(n_components=n)
    features = pca_n.fit_transform(features)
    model = LogisticRegression()
    cv_score = cross_val_score(model, features, y_train, cv=5, scoring='accuracy').mean()
    scores.append(cv_score)
plt.plot(range(1, pca.n_components_ + 1), scores)
plt.xlabel('Number of Components')
plt.ylabel('Cross-Validated Accuracy')
plt.show()

# PCA
'https://scikit-learn.org/1.5/modules/generated/sklearn.decomposition.PCA.html'
pca = PCA(n_components=10)  # Put determined number of components here
pca.fit(features)
print('Explained variance ratio:', pca.explained_variance_ratio_)
print('Feature contributions', pca.components_)

